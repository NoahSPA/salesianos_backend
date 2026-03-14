from __future__ import annotations

import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import Response
from openpyxl import load_workbook

from app.api.deps import get_current_user, require_roles
from app.core.validators import normalize_rut
from app.db.ids import oid
from app.domains.audit.service import log_audit
from app.domains.players.repo import create_player, get_blocked_dorsals, get_player, list_players, update_player, upsert_by_rut
from app.domains.players.schemas import PlayerCreate, PlayerImportResult, PlayerOut, PlayerUpdate
from app.domains.series.repo import get_series, get_series_by_name
from app.storage.gridfs import delete_avatar_file, get_avatar_file, upload_avatar

router = APIRouter(prefix="/players", tags=["players"])


def _ensure_primary_in_series(primary: str, series: list[str]) -> list[str]:
    if primary not in series:
        return [primary, *series]
    return series


# Mapeo de cabeceras Excel/CSV (español e inglés) a clave canónica
_HEADER_ALIASES: dict[str, str] = {
    "nombre": "first_name",
    "nombres": "first_name",
    "primer nombre": "first_name",
    "first_name": "first_name",
    "segundo nombre": "second_first_name",
    "second_first_name": "second_first_name",
    "apellido": "last_name",
    "apellidos": "last_name",
    "primer apellido": "last_name",
    "last_name": "last_name",
    "segundo apellido": "second_last_name",
    "second_last_name": "second_last_name",
    "rut": "rut",
    "run": "rut",
    "documento": "rut",
    "fecha_nacimiento": "birth_date",
    "fecha nacimiento": "birth_date",
    "fecha de nacimiento": "birth_date",
    "nacimiento": "birth_date",
    "birth_date": "birth_date",
    "telefono": "phone",
    "teléfono": "phone",
    "phone": "phone",
    "celular": "phone",
    "fono": "phone",
    "email": "email",
    "correo": "email",
    "serie_principal_id": "primary_series_id",
    "primary_series_id": "primary_series_id",
    "serie": "primary_series_id",
    "serie principal": "primary_series_id",
    "series_ids": "series_ids",
    "series": "series_ids",
    "posicion_principal": "position_primary",
    "posición principal": "position_primary",
    "position_primary": "position_primary",
    "posicion": "position_primary",
    "posición": "position_primary",
    "posicion principal": "position_primary",
    "posicion_secundaria": "position_secondary",
    "posición secundaria": "position_secondary",
    "posición segundaria": "position_secondary",
    "position_secondary": "position_secondary",
    "posicion secundaria": "position_secondary",
    "posicion segundaria": "position_secondary",
    "nivel": "level",
    "level": "level",
    "observaciones": "notes",
    "notes": "notes",
    "nota": "notes",
}


def _normalize_header(h: str) -> str:
    """Convierte cabecera a clave canónica para la fila. Quita tildes para matching."""
    s = (h or "").strip().lower().replace("  ", " ")
    # Sin tildes para que "Posicion" y "Posición" coincidan
    for old, new in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")]:
        s = s.replace(old, new)
    return _HEADER_ALIASES.get(s, s)


def _row_value(row: dict, *keys: str) -> str:
    """Obtiene el valor de la fila por clave canónica; acepta varias claves."""
    for k in keys:
        v = row.get(k)
        if v is None:
            continue
        if isinstance(v, (date, datetime)):
            return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)
        s = str(v).strip()
        if s:
            return s
    return ""


async def _process_import_row(
    row: dict,
    line: int,
    series_id_override: str | None = None,
) -> tuple[str, dict]:
    """
    Procesa una fila de importación (CSV o Excel ya normalizada).
    Si series_id_override está definido (carga por serie), se usa como serie principal.
    Devuelve (mode, player_out). Lanza excepción si hay error de validación.
    """
    first_name = _row_value(row, "first_name")
    second_first_name = _row_value(row, "second_first_name") or None
    last_name = _row_value(row, "last_name")
    second_last_name = _row_value(row, "second_last_name") or None
    rut_raw = _row_value(row, "rut")
    if not rut_raw:
        raise ValueError("RUT vacío")
    rut = normalize_rut(rut_raw)
    birth = _row_value(row, "birth_date")
    phone = _row_value(row, "phone")
    email = _row_value(row, "email") or None
    position_primary_raw = _row_value(row, "position_primary")
    position_secondary_raw = _row_value(row, "position_secondary")
    position_primary = position_primary_raw.strip() or ""
    position_secondary = (position_secondary_raw.strip() or None) if position_secondary_raw else None
    level = _row_value(row, "level")
    notes = _row_value(row, "notes") or None

    if series_id_override:
        primary_series_id = series_id_override
        series_ids = [primary_series_id]
        existing_series = await get_series(primary_series_id)
        if not existing_series:
            raise ValueError(f"La serie con ID {primary_series_id!r} no existe")
        resolved_primary_id = primary_series_id
    else:
        primary_series_id = _row_value(row, "primary_series_id")
        series_ids_raw = _row_value(row, "series_ids")
        if not (first_name and last_name and birth and primary_series_id and position_primary):
            raise ValueError("Faltan campos obligatorios (nombre, apellido, RUT, fecha nacimiento, serie, posición)")
        series_ids = [x.strip() for x in series_ids_raw.split(",") if x.strip()] if series_ids_raw else []
        existing_series = await get_series(primary_series_id)
        if not existing_series:
            existing_series = await get_series_by_name(primary_series_id)
        if not existing_series:
            raise ValueError(f"La serie '{primary_series_id}' no existe (use ID o nombre)")
        resolved_primary_id = existing_series["id"]
        series_ids = _ensure_primary_in_series(resolved_primary_id, series_ids)

    # En carga masiva (Excel/CSV) el celular no es obligatorio
    missing = []
    if not first_name:
        missing.append("Primer nombre")
    if not last_name:
        missing.append("Primer apellido")
    if not birth:
        missing.append("Fecha nacimiento")
    if missing:
        raise ValueError("Faltan campos obligatorios: " + ", ".join(missing))

    # En carga por serie, posiciones vacías en el Excel se aceptan: por defecto "cm"
    if series_id_override:
        if not position_primary:
            position_primary = "cm"
        if position_secondary is not None and not position_secondary.strip():
            position_secondary = None
    elif not position_primary:
        raise ValueError("Debe indicar al menos Posición Principal")

    # Aceptar AAAA-MM-DD o DD/MM/AAAA
    birth_date: date
    try:
        birth_date = date.fromisoformat(birth)
    except ValueError:
        parts = birth.strip().split("/")
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000 if y < 50 else 1900
            try:
                birth_date = date(y, m, d)
            except ValueError:
                raise ValueError(f"Fecha de nacimiento inválida: {birth!r}. Use AAAA-MM-DD o DD/MM/AAAA")
        else:
            raise ValueError(f"Fecha de nacimiento inválida: {birth!r}. Use AAAA-MM-DD o DD/MM/AAAA")

    positions = [x.strip() for x in [position_primary, position_secondary or ""] if x.strip()]
    if not positions:
        raise ValueError("Debe indicar al menos Posición Principal")

    if level:
        try:
            level_stars = int(level)
        except ValueError:
            lvl = level.strip().lower()
            level_stars = 2 if lvl == "bajo" else 3 if lvl == "medio" else 4 if lvl == "alto" else 3
        level_stars = max(1, min(5, level_stars))
    else:
        level_stars = 3

    doc = {
        "first_name": first_name,
        "last_name": last_name,
        "rut": rut,
        "birth_date": birth_date,
        "phone": phone or "",
        "primary_series_id": oid(resolved_primary_id),
        "series_ids": [oid(x) for x in series_ids],
        "positions": positions,
        "level_stars": level_stars,
        "active": True,
        "notes": notes,
    }
    if second_first_name is not None:
        doc["second_first_name"] = second_first_name
    if second_last_name is not None:
        doc["second_last_name"] = second_last_name
    if email is not None:
        doc["email"] = email

    return await upsert_by_rut(rut=rut, doc=doc)


@router.get("", response_model=list[PlayerOut], dependencies=[Depends(get_current_user)])
async def players_list(active: bool | None = None, series_id: str | None = None, q: str | None = None) -> list[PlayerOut]:
    docs = await list_players(active=active, series_id=series_id, q=q)
    return [PlayerOut(**d) for d in docs]


@router.post("", response_model=PlayerOut, dependencies=[Depends(require_roles("admin", "delegado"))])
async def players_create(payload: PlayerCreate, actor=Depends(get_current_user)) -> PlayerOut:
    data = payload.model_dump()
    raw_primary = (data.get("primary_series_id") or "").strip()
    if raw_primary:
        try:
            data["primary_series_id"] = oid(raw_primary)
            data["series_ids"] = [oid(x) for x in _ensure_primary_in_series(payload.primary_series_id, payload.series_ids)]
        except ValueError:
            raise HTTPException(status_code=400, detail="Serie principal inválida")
    else:
        if not payload.in_memoriam:
            raise HTTPException(status_code=400, detail="Falta serie principal")
        data["primary_series_id"] = None
        data["series_ids"] = []
    # Dorsales de jugadores en memoria están bloqueados para otros
    if not payload.in_memoriam and payload.dorsal is not None:
        blocked = await get_blocked_dorsals()
        if payload.dorsal in blocked:
            raise HTTPException(
                status_code=400,
                detail=f"El dorsal {payload.dorsal} está reservado en memoria y no puede asignarse a otro jugador.",
            )
    # Mongo solo acepta tipos simples: positions como list[str]
    data["positions"] = [p.value if hasattr(p, "value") else str(p) for p in (data.get("positions") or [])]
    data.pop("position_primary", None)
    data.pop("position_secondary", None)
    data.pop("level", None)
    created = await create_player(data)
    await log_audit(
        actor=actor,
        action="player_created",
        entity_type="player",
        entity_id=created["id"],
        after=created,
    )
    return PlayerOut(**created)


@router.get("/{player_id}", response_model=PlayerOut, dependencies=[Depends(get_current_user)])
async def players_get(player_id: str) -> PlayerOut:
    doc = await get_player(player_id)
    if not doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No encontrado")
    return PlayerOut(**doc)


@router.get("/{player_id}/avatar")
async def players_get_avatar(player_id: str) -> Response:
    """Sirve el avatar del jugador desde GridFS. Si no tiene avatar_file_id, 404."""
    doc = await get_player(player_id)
    if not doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No encontrado")
    file_id = doc.get("avatar_file_id")
    if not file_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No tiene avatar")
    result = await get_avatar_file(file_id)
    if not result:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Avatar no encontrado")
    data, content_type = result
    return Response(content=data, media_type=content_type)


@router.post("/{player_id}/avatar", response_model=PlayerOut, dependencies=[Depends(require_roles("admin", "delegado"))])
async def players_upload_avatar(
    player_id: str,
    actor=Depends(get_current_user),
    file: UploadFile = File(...),
) -> PlayerOut:
    """Sube el avatar del jugador. Acepta image/jpeg, image/png, image/webp. Comprime si supera el límite."""
    doc = await get_player(player_id)
    if not doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No encontrado")
    ct = file.content_type or ""
    if ct not in {"image/jpeg", "image/png", "image/webp", "image/jpg"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato no soportado. Use JPEG, PNG o WebP",
        )
    if ct == "image/jpg":
        ct = "image/jpeg"
    raw = await file.read()
    if len(raw) < 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Archivo vacío")
    old_file_id = doc.get("avatar_file_id")
    try:
        file_id = await upload_avatar(raw, ct, filename=file.filename or "avatar")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    await update_player(player_id, {"avatar_file_id": file_id})
    if old_file_id:
        await delete_avatar_file(old_file_id)
    out = await get_player(player_id)
    await log_audit(
        actor=actor,
        action="player_avatar_uploaded",
        entity_type="player",
        entity_id=player_id,
        after={"avatar_file_id": file_id},
    )
    return PlayerOut(**out)


@router.patch("/{player_id}", response_model=PlayerOut, dependencies=[Depends(require_roles("admin", "delegado"))])
async def players_patch(player_id: str, payload: PlayerUpdate, actor=Depends(get_current_user)) -> PlayerOut:
    before = await get_player(player_id)
    if not before:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No encontrado")
    patch = payload.model_dump(exclude_unset=True)
    if patch.get("primary_series_id") and patch.get("series_ids") is not None:
        patch["series_ids"] = _ensure_primary_in_series(patch["primary_series_id"], patch["series_ids"])
    if "positions" in patch and patch["positions"] is not None:
        patch["positions"] = [p.value if hasattr(p, "value") else str(p) for p in patch["positions"]]
    patch.pop("position_primary", None)
    patch.pop("position_secondary", None)
    patch.pop("level", None)
    # Dorsales de jugadores en memoria están bloqueados para otros
    final_in_memoriam = patch.get("in_memoriam") if "in_memoriam" in patch else before.get("in_memoriam", False)
    final_dorsal = patch.get("dorsal") if "dorsal" in patch else before.get("dorsal")
    if not final_in_memoriam and final_dorsal is not None:
        blocked = await get_blocked_dorsals(exclude_player_id=player_id)
        if final_dorsal in blocked:
            raise HTTPException(
                status_code=400,
                detail=f"El dorsal {final_dorsal} está reservado en memoria y no puede asignarse a otro jugador.",
            )
    after = await update_player(player_id, patch)
    await log_audit(
        actor=actor,
        action="player_updated",
        entity_type="player",
        entity_id=player_id,
        before=before,
        after=after,
    )
    return PlayerOut(**after)


@router.post("/import-csv", response_model=PlayerImportResult, dependencies=[Depends(require_roles("admin", "delegado"))])
async def players_import_csv(
    actor=Depends(get_current_user),
    file: UploadFile = File(...),
) -> PlayerImportResult:
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Archivo debe ser CSV")
    raw = await file.read()
    if len(raw) > 2_000_000:
        raise HTTPException(status_code=400, detail="CSV demasiado grande (máx 2MB)")

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    inserted = updated = skipped = 0
    errors: list[dict] = []

    for idx, row in enumerate(reader, start=2):  # header=1
        row_canon = {_normalize_header(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
        try:
            mode, out = await _process_import_row(row_canon, idx)
            if mode == "inserted":
                inserted += 1
                await log_audit(actor=actor, action="player_created_csv", entity_type="player", entity_id=out["id"], after=out)
            else:
                updated += 1
                await log_audit(actor=actor, action="player_updated_csv", entity_type="player", entity_id=out["id"], after=out)
        except Exception as e:
            skipped += 1
            errors.append({"line": idx, "error": str(e), "row": {k: str(v)[:50] if v is not None else "" for k, v in row.items()}})

    return PlayerImportResult(inserted=inserted, updated=updated, skipped=skipped, errors=errors)


@router.post("/import-excel", response_model=PlayerImportResult, dependencies=[Depends(require_roles("admin", "delegado"))])
async def players_import_excel(
    actor=Depends(get_current_user),
    series_id: str = Form(..., description="ID de la serie a la que pertenecen los jugadores de la nómina"),
    file: UploadFile = File(...),
) -> PlayerImportResult:
    """
    Carga masiva de jugadores desde un Excel (.xlsx) por serie.
    La nómina se sube para una serie concreta (series_id obligatorio).
    Primera fila: cabeceras (RUT, Primer Nombre, Segundo Nombre, Primer Apellido, Segundo Apellido,
    Fecha Nacimiento, Celular, Email, Posición Principal, Posición Secundaria).
    RUT puede venir con o sin puntos. Si el RUT ya existe, se actualiza (no se duplica).
    """
    series_id = (series_id or "").strip()
    if not series_id:
        raise HTTPException(status_code=400, detail="Debe indicar la serie (series_id)")
    existing_series = await get_series(series_id)
    if not existing_series:
        raise HTTPException(status_code=400, detail="La serie indicada no existe")

    if not file.filename or not (file.filename.lower().endswith(".xlsx") or file.filename.lower().endswith(".xls")):
        raise HTTPException(status_code=400, detail="Archivo debe ser Excel (.xlsx)")
    raw = await file.read()
    if len(raw) > 5_000_000:
        raise HTTPException(status_code=400, detail="Excel demasiado grande (máx 5MB)")

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    # Usar hoja "Jugadores" si existe, si no la activa
    sheet_name = "jugadores"
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="El libro no tiene hojas. Debe existir una hoja llamada 'Jugadores' o tener datos en la primera hoja.")

    inserted = updated = skipped = 0
    errors: list[dict] = []

    # Primera fila = cabeceras (normalizar a claves canónicas)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="La primera fila debe contener las cabeceras")
    headers = [_normalize_header(str(c or "").strip()) for c in header_row]

    def _cell_value(v):  # Normalizar valor de celda: números (ej. RUT) a string, strip
        if v is None:
            return None
        if isinstance(v, (date, datetime)):
            return v
        s = str(v).strip()
        return s if s else None

    for row_idx, row_tuple in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row_tuple):
            continue
        row_canon = {}
        for col_idx, key in enumerate(headers):
            if not key:
                continue
            val = row_tuple[col_idx] if col_idx < len(row_tuple) else None
            normalized = _cell_value(val)
            row_canon[key] = normalized if normalized is not None else ""
        try:
            mode, out = await _process_import_row(row_canon, row_idx, series_id_override=series_id)
            if mode == "inserted":
                inserted += 1
                await log_audit(actor=actor, action="player_created_excel", entity_type="player", entity_id=out["id"], after=out)
            else:
                updated += 1
                await log_audit(actor=actor, action="player_updated_excel", entity_type="player", entity_id=out["id"], after=out)
        except Exception as e:
            skipped += 1
            errors.append({
                "line": row_idx,
                "error": str(e),
                "row": {k: str(v)[:50] if v is not None else "" for k, v in row_canon.items()},
            })

    wb.close()
    return PlayerImportResult(inserted=inserted, updated=updated, skipped=skipped, errors=errors)

